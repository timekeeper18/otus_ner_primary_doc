import requests
from service_settings import microservices
import json
import ast
import pandas as pd
from natasha import (
    NewsEmbedding,
    NewsNERTagger)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re


def get_tag_deeppavlov(txt, tag):
    body = dict(txt=txt, tag=tag)

    # received_data = requests.post('http://{}:{}/deeppavlov'.format("192.168.18.194", 8080), json=body)

    received_data = requests.post(microservices['DEEPPAVLOV']['url'], json=body,
                                  headers={'Content-Type': 'application/json', "charset": "utf-8"})
    job_result = json.loads(received_data.text).get('entities', [])
    # self.entity_queue.send_job('deeppavlov_entity.get_entities_dp', params)
    # return self.entity_queue.get_job_result()
    return job_result

def get_tag_natasha(txt):
    emb = NewsEmbedding()
    ner_tagger = NewsNERTagger(emb)

    NER = ner_tagger(txt)
    n_ner = len(NER.spans)
    tag_list = []
    if n_ner > 0:
        for i in range(n_ner):
            t = NER.spans[i].type
            start = NER.spans[i].start
            stop = NER.spans[i].stop
            tag_list.append([txt[start:stop], t])
    return tag_list

def get_quality_deeppavlov(input_string, correct_result):
    result_tag = get_tag_deeppavlov(input_string, 'ORG')
    result_tag = [i for i in result_tag if 'ORG' in i or 'PER' in i]
    result_tag = [i[::-1] for i in result_tag]


def get_quality_natasha(input_string, correct_result):
    result_tag = get_tag_natasha(input_string)
    result_tag = [i for i in result_tag if 'ORG' in i or 'PER' in i]
    result_tag = [i[::-1] for i in result_tag]

    return result_tag


def get_string_and_tag_value(input_df):
    """
    Функция получения текста и правильного ответа по тэгам и передачи в функцию сравнения DeepPavlov and Natasha
    """
    redFill = PatternFill(start_color='00FF0000',
                          end_color='00FF0000',
                          fill_type='solid')
    grnFill = PatternFill(start_color='0000FF00',
                          end_color='0000FF00',
                          fill_type='solid')
    yelFill = PatternFill(start_color='00FFFF00',
                          end_color='0000FF00',
                          fill_type='solid')
    wb = load_workbook("/opt/dev/NER_TEST/NER_TEST.xlsx")
    ws = wb.active
    full_correct_amount = 0
    part_correct_amount = 0
    df_ner_text = df_ner['text']
    df_ner_teg_value = df_ner['teg_value']
    len_df = input_df.shape[0]
    count_org_tag = 0
    count_correct_dp_all = 0
    count_correct_nt_all = 0
    for row_count in range(len_df): 
        text = df_ner_text[row_count]  # Исходная строка
        teg_value = df_ner_teg_value[row_count]
        if isinstance(teg_value, str):  # В excel есть записи list , есть str -> "[['ORG', 'ООО РОМАШКА']]"
            teg_value = ast.literal_eval(teg_value)  # Преобразование из str в list
        teg_value_start = teg_value
        teg_value = [i for i in teg_value if 'ORG' in i or 'PER' in i]

        result_tag_dp = get_quality_deeppavlov(text, teg_value)  # Сравнение резов DeepPavlov и правильных
        print(result_tag_dp, text)
        result_tag_nt = get_quality_natasha(text, teg_value)

        ws['A{}'.format(row_count+2)] = row_count
        ws['B{}'.format(row_count+2)] = text      # Исходная строка
        ws['C{}'.format(row_count+2)] = str(teg_value_start) # Правильные тэги
        ws['D{}'.format(row_count+2)] = str(result_tag_dp) # результат DeppPavlov
        ws['E{}'.format(row_count+2)] = str(result_tag_nt) # результат Natasha

        count_correct_dp = 0
        count_correct_nt = 0

        for org_tag in teg_value:
            org_tag = [''.join(re.findall('[A-Za-zА-Яа-я0-9]', i)).lower().replace(' ', '') for i in org_tag]
            new_result_tag_dp = []
            for i in result_tag_dp:
                i = [''.join(re.findall('[A-Za-zА-Яа-я0-9]', f)).lower().replace(' ', '') for f in i]
                new_result_tag_dp.append(i)

            if org_tag in new_result_tag_dp:
                count_correct_dp += 1
                count_correct_dp_all += 1

            new_result_tag_nt = []
            for i in result_tag_nt:
                i = [''.join(re.findall('[A-Za-zА-Яа-я0-9]', f)).lower().replace(' ', '') for f in i]
                new_result_tag_nt.append(i)
            if org_tag in new_result_tag_nt:
                count_correct_nt += 1
                count_correct_nt_all += 1
            count_org_tag += 1
            if count_correct_nt == 0:
                ws['E{}'.format(row_count+2)].fill = redFill
            if count_correct_nt == len(teg_value) and count_correct_nt == len(result_tag_nt):
                ws['E{}'.format(row_count+2)].fill = grnFill
            if (count_correct_nt < len(teg_value) or len(result_tag_nt) > len(teg_value)) and count_correct_nt != 0:
                ws['E{}'.format(row_count+2)].fill = yelFill

            if count_correct_dp == 0:
                ws['D{}'.format(row_count+2)].fill = redFill
            if count_correct_dp == len(teg_value) and count_correct_dp == len(result_tag_dp):
                ws['D{}'.format(row_count+2)].fill = grnFill
            if (count_correct_dp < len(teg_value) or len(result_tag_dp) > len(teg_value)) and count_correct_dp != 0:
                ws['D{}'.format(row_count+2)].fill = yelFill

    wb.save("/opt/dev/NER_TEST/NER_TEST.xlsx")
    return count_org_tag, count_correct_dp_all, count_correct_nt_all

if __name__ == "__main__":
    df_ner = pd.read_excel("/opt/dev/NER_TEST/test_model_example.xlsx")
    correct_result = get_string_and_tag_value(df_ner)
    print(correct_result)